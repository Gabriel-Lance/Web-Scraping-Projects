"""A web scraper that scrapes the title, genre and price of every book on books.toscrape.com
and puts them in an Excel workbook"""

import requests
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook

# Gets the links to every book genre
url = "http://books.toscrape.com/catalogue/category/books_1/index.html"
main_page = bs(requests.get(url).text, "html.parser")
genres = main_page.select("ul>li>ul>li>a")
genre_links = ["http://books.toscrape.com/catalogue/category" + genre['href'][2:] for genre in genres]

# Creates an Excel workbook to save the data
wb = Workbook()
ws = wb.active
ws["A1"].value = "Title"
ws["B1"].value = "Genre"
ws["C1"].value = "Price"
current_row = 2

# Begins scraping
for genre_link in genre_links:
    # Downloads the page for the next genre
    current_page = bs(requests.get(genre_link).text, "html.parser")
    current_genre_name = current_page.find("h1").text
    print(f"\nCurrent genre: {current_genre_name}\n")

    # Gets links for every book in the current genre
    book_links = ["http://books.toscrape.com/catalogue" + book["href"][8:] for book in current_page.select("h3>a")]

    # Scrapes every page in the list of links
    for book_link in book_links:
        current_book = bs(requests.get(book_link).text, "html.parser")
        current_book_title = current_book.find("h1").text
        current_book_price = current_book.select("p.price_color")[0].text[1:]
        print(f"Adding \"{current_book_title}\"")
        ws.cell(row=current_row, column=1).value = current_book_title
        ws.cell(row=current_row, column=2).value = current_genre_name
        ws.cell(row=current_row, column=3).value = current_book_price
        current_row += 1

# Saves the workbook
wb.save("book_scrape.xlsx")
print("Done!")

