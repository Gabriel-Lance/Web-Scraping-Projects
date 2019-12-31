"""A web scraper that scrapes a list of quotes, along with each quote's author and tags, from quotes.toscrape.com.
The scraped data is then saved to an Excel workbook.
This is a simple web page with straightforward HTML, making it easy to scrape."""

import requests
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook

url = "http://quotes.toscrape.com"

# Creates the workbook to save the data
wb = Workbook()
ws = wb.active
ws["A1"].value = "Quote"
ws["B1"].value = "Author"
ws["C1"].value = "Tags"
current_row = 2

# Begins scraping
while True:
    # Downloads the next page to scrape
    current_page = bs(requests.get(url).text, "html.parser")
    quote_blocks = current_page.select("div[class='quote']")

    # Scrapes the page
    for quote_block in quote_blocks:
        quote = quote_block.select("span[class='text']")[0].text
        author = quote_block.select("span > small[class='author']")[0].text
        tags = [tag.text for tag in quote_block.select("a[class='tag']")]
        ws.cell(row=current_row, column=1).value = quote
        ws.cell(row=current_row, column=2).value = author
        ws.cell(row=current_row, column=3).value = ", ".join(tags)
        print(f"{current_row - 1} quotes scraped so far.")
        current_row += 1

    # Finds the next URL to scrape, or stops scraping if this is the last page
    try:
        url = "http://quotes.toscrape.com" + current_page.select("li[class='next']>a")[0]["href"]
    except IndexError:
        break

wb.save("quote_scrape.xlsx")
print("Done!")
