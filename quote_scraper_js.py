"""A web scraper that scrapes a list of quotes, along with each quote's author and list of tags, from
quotes.toscrape.com/js.  The HTML on this page is initially empty and is populated by JavaScript, which means that
it cannot be scraped using Requests and BeautifulSoup, as those libraries don't execute JavaScript.  Thus, this
scraper uses Selenium to control Chrome, which will execute the required JavaScript.

To run this code, you need to have an appropriate version of the Chrome webdriver in the same directory as this file."""

from openpyxl import Workbook
from selenium import webdriver

# Creates an Excel workbook to save the data
wb = Workbook()
ws = wb.active
ws["A1"].value = "Quote"
ws["B1"].value = "Author"
ws["C1"].value = "Tags"

# Opens the first page
url = "http://quotes.toscrape.com/js/"
browser = webdriver.Chrome()

# Begins scraping
current_row = 2
browser.get(url)
while True:
    quote_blocks = [block.text.split("\n") for block in browser.find_elements_by_class_name("quote")]
    for quote_block in quote_blocks:
        quote = quote_block[0]
        author = quote_block[1][quote_block[1].index(" ")+1:]
        tags = ", ".join(quote_block[2].split()[1:])

        ws.cell(row=current_row, column=1).value = quote
        ws.cell(row=current_row, column=2).value = author
        ws.cell(row=current_row, column=3).value = tags
        current_row += 1
    try:
        browser.find_element_by_css_selector("li[class='next'] > a").click()
    except:
        break

# Saves the workbook
wb.save("quote_scrape_JS.xlsx")
print("DONE!")
