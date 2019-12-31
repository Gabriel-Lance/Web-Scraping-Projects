"""A web scraper that gets every quote, author, and group of tags from quotes.toscrape.com/tableful and saves the
data to an Excel workbook. The HTML of this webpage is formatted in a strange way, where the page is a table and each
cell of the table alternates between containing a quote and its author and containing a quote's list of tags.  This
makes it difficult to scrape."""

import requests
from bs4 import BeautifulSoup as bs
import re
from openpyxl import Workbook

# Creates a workbook to save the data
wb = Workbook()
ws = wb.active
ws["A1"].value = "Quote"
ws["B1"].value = "Author"
ws["C1"].value = "Tags"

# Gets a list of all pages that need to be visited
page = requests.get("http://quotes.toscrape.com/tableful/").text
souped = bs(page, "html.parser")
links = souped.select("table > tr > td")[1].select("a")
urls = []
for link in links:
    urls.append("http://quotes.toscrape.com" + link["href"])

# Scrapes each page
current_row = 2
quote_finder = re.compile(r'“.*”')
for url in urls:
    page = requests.get(url)
    souped = bs(page.text, "html.parser")
    elems = souped.select("table > tr > td")[2:]

    for elem in elems:
        if "Tags" not in elem.text:
            quote = quote_finder.search(elem.text)
            if quote is None:
                break
            quote = quote.group()
            author = elem.text.split("Author: ")[1]
            ws.cell(row=current_row, column=1).value = quote
            ws.cell(row=current_row, column=2).value = author
        else:
            tags = [tag.text for tag in elem.select("a")]
            ws.cell(row=current_row, column=3).value = ", ".join(tags)
            current_row += 1

# Saves the workbook
wb.save("quote_scrape_table.xlsx")
print("Done!")
