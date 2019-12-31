"""A web scraper that scrapes a list of quotes, along with each quote's author and list of tags, from
quotes.toscrape.com/scroll.  Here, the HTML doesn't contain all quotes when the page is first loaded - the user must
scroll to the bottom of the page before more quotes are added. To deal with this, this scraper uses Selenium to control
Google Chrome, so it can scroll to the bottom of the page to reveal every quote.

To run this code, you need to have an appropriate version of the Chrome webdriver in the same directory as this file."""

from openpyxl import Workbook
from selenium import webdriver
from time import sleep

# Creates an Excel workbook to save the data
wb = Workbook()
ws = wb.active
ws["A1"].value = "Quote"
ws["B1"].value = "Author"
ws["C1"].value = "Tags"

# Opens the page
browser = webdriver.Chrome()
browser.get("http://quotes.toscrape.com/scroll")
browser.implicitly_wait(1)


# Scrolls to the bottom of the page
quote_count = len(browser.find_elements_by_class_name("quote"))
while True:
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    sleep(1)
    new_quote_count = len(browser.find_elements_by_class_name("quote"))
    if new_quote_count == quote_count:
        break
    else:
        quote_count = new_quote_count

# Begins scraping
quotes = browser.find_elements_by_css_selector("div[class='quote'] > span[class='text']")
authors = browser.find_elements_by_css_selector("span > small[class='author']")
tags = [", ".join(element.text.split()[1:]) for element in browser.find_elements_by_class_name("tags")]
current_row = 2
for i in range(len(quotes)):
    ws.cell(row=current_row, column=1).value = quotes[i].text
    ws.cell(row=current_row, column=2).value = authors[i].text
    ws.cell(row=current_row, column=3).value = tags[i]
    current_row += 1

# Saves the workbook
wb.save("quote_scrape_2.xlsx")
print("DONE!")